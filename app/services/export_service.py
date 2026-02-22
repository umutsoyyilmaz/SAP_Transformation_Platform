import csv
import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAG_FILLS = {
    "green": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "amber": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
    "red": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
}
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill(start_color="354A5F", end_color="354A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_program_health_xlsx(health_data: dict) -> io.BytesIO:
    """
    Generate a styled Excel workbook from program health data.
    Returns a BytesIO buffer ready for Flask send_file.
    """
    wb = Workbook()

    # ── Sheet 1: Executive Summary ────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Program Health Report — {health_data['program_name']}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    if health_data.get("days_to_go_live") is not None:
        ws["D2"] = f"Days to Go-Live: {health_data['days_to_go_live']}"
        ws["D2"].font = Font(size=11, bold=True)

    # Overall RAG
    ws["A4"] = "Overall Status"
    ws["A4"].font = Font(size=12, bold=True)
    ws["B4"] = health_data["overall_rag"].upper()
    ws["B4"].fill = RAG_FILLS.get(health_data["overall_rag"], RAG_FILLS["amber"])
    ws["B4"].font = WHITE_FONT
    ws["B4"].alignment = Alignment(horizontal="center")

    # Area breakdown table
    row = 6
    headers = ["Area", "RAG", "Key Metric", "Value", "Details"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    areas = health_data.get("areas", {})
    area_rows = [
        (
            "Explore",
            areas.get("explore", {}),
            "Workshop Completion",
            f"{areas.get('explore', {}).get('workshops', {}).get('pct', 0)}%",
            f"{areas.get('explore', {}).get('requirements', {}).get('total', 0)} reqs, "
            f"{areas.get('explore', {}).get('open_items', {}).get('overdue', 0)} overdue OIs",
        ),
        (
            "Backlog",
            areas.get("backlog", {}),
            "Completion",
            f"{areas.get('backlog', {}).get('items', {}).get('pct', 0)}%",
            f"{areas.get('backlog', {}).get('items', {}).get('total', 0)} total items",
        ),
        (
            "Testing",
            areas.get("testing", {}),
            "Pass Rate",
            f"{areas.get('testing', {}).get('pass_rate', 0)}%",
            f"{areas.get('testing', {}).get('defects', {}).get('open', 0)} open defects, "
            f"{areas.get('testing', {}).get('defects', {}).get('s1_open', 0)} S1",
        ),
        (
            "RAID",
            areas.get("raid", {}),
            "Open Risks",
            f"{areas.get('raid', {}).get('risks_open', 0)}",
            f"{areas.get('raid', {}).get('risks_red', 0)} red risks, "
            f"{areas.get('raid', {}).get('actions_overdue', 0)} overdue actions",
        ),
        (
            "Integration",
            areas.get("integration", {}),
            "Live %",
            f"{areas.get('integration', {}).get('interfaces', {}).get('pct', 0)}%",
            f"{areas.get('integration', {}).get('interfaces', {}).get('total', 0)} total interfaces",
        ),
    ]

    for area_name, area_data, metric, value, details in area_rows:
        row += 1
        ws.cell(row=row, column=1, value=area_name).border = THIN_BORDER
        rag_cell = ws.cell(row=row, column=2, value=area_data.get("rag", "—").upper())
        rag_cell.fill = RAG_FILLS.get(area_data.get("rag", ""), PatternFill())
        rag_cell.font = WHITE_FONT
        rag_cell.alignment = Alignment(horizontal="center")
        rag_cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=4, value=value).border = THIN_BORDER
        ws.cell(row=row, column=5, value=details).border = THIN_BORDER

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = [20, 10, 22, 12, 45][col - 1]

    # ── Sheet 2: Phase Timeline ───────────────────────────────────────
    ws2 = wb.create_sheet("Phases")
    headers2 = ["Phase", "Status", "Completion %", "Planned Start", "Planned End"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for i, phase in enumerate(health_data.get("phases", []), 2):
        ws2.cell(row=i, column=1, value=phase["name"]).border = THIN_BORDER
        ws2.cell(row=i, column=2, value=phase["status"]).border = THIN_BORDER
        pct_cell = ws2.cell(row=i, column=3, value=phase["completion_pct"])
        pct_cell.border = THIN_BORDER
        pct_cell.alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=4, value=phase.get("planned_start", "")).border = THIN_BORDER
        ws2.cell(row=i, column=5, value=phase.get("planned_end", "")).border = THIN_BORDER

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = [25, 12, 14, 14, 14][col - 1]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_program_health_html(health_data: dict) -> str:
    """
    Generate an HTML report suitable for printing / PDF conversion.
    Returns HTML string with inline CSS for print-friendliness.
    """
    rag_colors = {"green": "#27AE60", "amber": "#F39C12", "red": "#E74C3C"}
    areas = health_data.get("areas", {})

    area_rows_html = ""
    for name, key in [
        ("Explore", "explore"),
        ("Backlog", "backlog"),
        ("Testing", "testing"),
        ("RAID", "raid"),
        ("Integration", "integration"),
    ]:
        area = areas.get(key, {})
        rag = area.get("rag", "—")
        color = rag_colors.get(rag, "#999")
        area_rows_html += f"""
        <tr>
            <td style=\"font-weight:600\">{name}</td>
            <td style=\"background:{color};color:#fff;text-align:center;font-weight:700;border-radius:4px\">{rag.upper()}</td>
        </tr>"""

    phases_html = ""
    for phase in health_data.get("phases", []):
        phases_html += f"""
        <tr>
            <td>{phase['name']}</td>
            <td>{phase['status']}</td>
            <td style=\"text-align:center\">{phase['completion_pct']}%</td>
            <td>{phase.get('planned_start', '')}</td>
            <td>{phase.get('planned_end', '')}</td>
        </tr>"""

    overall_rag = health_data.get("overall_rag", "—")
    overall_color = rag_colors.get(overall_rag, "#999")

    html = f"""<!DOCTYPE html>
<html><head>
<meta charset=\"utf-8\">
<title>Program Health Report — {health_data['program_name']}</title>
<style>
    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; }}
    h1 {{ color: #354A5F; margin-bottom: 4px; }}
    .meta {{ color: #666; font-size: 13px; margin-bottom: 24px; }}
    .overall {{ display: inline-block; padding: 8px 24px; background: {overall_color};
                color: #fff; font-weight: 700; font-size: 18px; border-radius: 6px; margin: 12px 0; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
    th {{ background: #354A5F; color: #fff; padding: 10px 12px; text-align: left; }}
    td {{ padding: 8px 12px; border-bottom: 1px solid #e0e0e0; }}
    tr:hover {{ background: #f5f7fa; }}
    @media print {{ body {{ margin: 20px; }} }}
</style>
</head><body>
<h1>Program Health Report</h1>
<p class=\"meta\">{health_data['program_name']} — Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}
{f" — Go-Live in {health_data['days_to_go_live']} days" if health_data.get('days_to_go_live') else ""}</p>

<div class=\"overall\">{overall_rag.upper()}</div>

<h2>Area Health</h2>
<table><thead><tr><th>Area</th><th style=\"width:100px\">RAG</th></tr></thead>
<tbody>{area_rows_html}</tbody></table>

<h2>Phase Timeline</h2>
<table><thead><tr><th>Phase</th><th>Status</th><th>Completion</th><th>Start</th><th>End</th></tr></thead>
<tbody>{phases_html}</tbody></table>

</body></html>"""

    return html


# ══════════════════════════════════════════════════════════════════════════════
# S2-04 (F-03) — Fit-Gap Report Export
#
# Audit A1: PDF dropped — only Excel + CSV (reviewer decision: weasyprint has
#   system-level deps that break Railway/Docker builds).
# Audit A2: Sync export is acceptable for ≤200 requirements. Async task queue
#   is a Phase 3 item when the 200+ threshold is routinely hit.
# Audit A3: Tenant isolation — all queries are scoped by tenant_id.
#   Export does NOT write temp files; returns in-memory bytes to avoid
#   cross-tenant path collisions.
# Audit A4: No "SAP standard template" branding — awaiting legal clearance.
# ══════════════════════════════════════════════════════════════════════════════

logger = logging.getLogger(__name__)


def _apply_header_style(ws, row: int, col_count: int) -> None:
    """Apply dark-header styling to an Excel row (1-indexed)."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    """Auto-size column widths based on content (capped at 60 chars)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_fitgap_excel(
    project_id: int,
    tenant_id: int | None,
    include_wricef: bool = True,
    include_config: bool = True,
    classification_filter: list[str] | None = None,
    sap_module_filter: list[str] | None = None,
    workshop_id: int | None = None,
) -> bytes:
    """Generate a Fit-Gap report Excel workbook (.xlsx) for a project.

    Produces 5 worksheets:
        1. Executive Summary  — headline counts and WRICEF breakdown.
        2. L3 Process Summary — per L3-process fit/partial/gap table.
        3. Requirement Detail — full list with classification, priority, status.
        4. WRICEF List        — BacklogItems linked to gap/partial_fit requirements.
        5. Config Items       — ConfigItems linked to fit requirements.

    Tenant isolation (Audit A3): all ORM queries are scoped by tenant_id.
    No temp files written — bytes returned in-memory.

    Args:
        project_id: Owning project (program_id / project_id).
        tenant_id: Row-level isolation. None = test environment.
        include_wricef: Include WRICEF tab (Tab 4). Default True.
        include_config: Include Config Items tab (Tab 5). Default True.
        classification_filter: Optional list of fit_status values to include.
        sap_module_filter: Optional list of SAP module codes to include.
        workshop_id: When set, restrict to requirements linked to this workshop.

    Returns:
        bytes: Raw .xlsx file content ready to stream to the client.
    """
    from app.models.explore import ExploreRequirement, ProcessLevel, ExploreWorkshop
    from app.models.backlog import BacklogItem, ConfigItem

    # ── Query requirements ────────────────────────────────────────────────────
    q = ExploreRequirement.query.filter(
        ExploreRequirement.project_id == project_id,
    )
    if tenant_id is not None:
        q = q.filter(ExploreRequirement.tenant_id == tenant_id)
    if classification_filter:
        q = q.filter(ExploreRequirement.fit_status.in_(classification_filter))
    if sap_module_filter:
        q = q.filter(ExploreRequirement.sap_module.in_(sap_module_filter))
    if workshop_id is not None:
        q = q.filter(ExploreRequirement.workshop_id == workshop_id)

    reqs = q.order_by(ExploreRequirement.code).all()

    # Aggregate counts by fit_status
    req_ids = [r.id for r in reqs]
    by_cls: dict[str, int] = {"fit": 0, "partial_fit": 0, "gap": 0}
    for r in reqs:
        cls = (r.fit_status or "gap").lower()
        by_cls[cls] = by_cls.get(cls, 0) + 1

    # ── Query WRICEF (BacklogItems linked to requirements in scope) ────────────
    wricef_items: list[BacklogItem] = []
    if include_wricef and req_ids:
        bi_q = BacklogItem.query.filter(
            BacklogItem.explore_requirement_id.in_(req_ids)
        )
        if tenant_id is not None:
            bi_q = bi_q.filter(BacklogItem.tenant_id == tenant_id)
        wricef_items = bi_q.order_by(BacklogItem.code).all()

    wricef_type_counts: dict[str, int] = {}
    for bi in wricef_items:
        wtype = (bi.wricef_type or "unknown").capitalize()
        wricef_type_counts[wtype] = wricef_type_counts.get(wtype, 0) + 1

    # ── Query Config Items ────────────────────────────────────────────────────
    config_items: list[ConfigItem] = []
    if include_config and req_ids:
        ci_q = ConfigItem.query.filter(
            ConfigItem.explore_requirement_id.in_(req_ids)
        )
        if tenant_id is not None:
            ci_q = ci_q.filter(ConfigItem.tenant_id == tenant_id)
        config_items = ci_q.order_by(ConfigItem.code).all()

    # ── Build process-level cache for Tab 2 ──────────────────────────────────
    pl_cache: dict[int, ProcessLevel] = {}
    for r in reqs:
        if r.process_level_id and r.process_level_id not in pl_cache:
            pl = db.session.get(ProcessLevel, r.process_level_id)
            if pl:
                pl_cache[r.process_level_id] = pl

    def _pl_ancestors(pl_id: int | None) -> dict[int, str]:
        """Return {level: name} for all ancestors of given ProcessLevel id."""
        result: dict[int, str] = {}
        visited: set[int] = set()
        current_id = pl_id
        while current_id and current_id not in visited:
            pl = pl_cache.get(current_id) or db.session.get(ProcessLevel, current_id)
            if not pl:
                break
            pl_cache[current_id] = pl
            visited.add(current_id)
            level = getattr(pl, "level", 0)
            result[level] = f"{pl.code} {pl.name}".strip()
            current_id = pl.parent_id
        return result

    # ── Build ProcessLevel→L3 group for Tab 2 ────────────────────────────────
    # Group requirements by their L3 process node
    l3_groups: dict[str, dict] = {}  # l3_key → {"l1","l2","l3", counts}
    for r in reqs:
        ancestors = _pl_ancestors(r.process_level_id)
        l3 = ancestors.get(3, "— Unknown —")
        l2 = ancestors.get(2, "")
        l1 = ancestors.get(1, "")
        key = l3
        if key not in l3_groups:
            l3_groups[key] = {
                "l1": l1, "l2": l2, "l3": l3,
                "fit": 0, "partial_fit": 0, "gap": 0, "total": 0,
            }
        cls = (r.fit_status or "gap").lower()
        l3_groups[key][cls] = l3_groups[key].get(cls, 0) + 1
        l3_groups[key]["total"] += 1

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"

    ws1.merge_cells("A1:C1")
    ws1["A1"] = "Fit-Gap Analysis Report"
    ws1["A1"].font = Font(size=16, bold=True, color="354A5F")
    ws1["A2"] = f"Project ID: {project_id}"
    ws1["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A3"].font = Font(italic=True, color="666666")

    summary_rows = [
        ("Field", "Count", ""),
        ("Total Requirements", len(reqs), ""),
        ("Fit", by_cls.get("fit", 0), f"{round(by_cls.get('fit',0)/max(len(reqs),1)*100,1)}%"),
        ("Partial Fit", by_cls.get("partial_fit", 0), f"{round(by_cls.get('partial_fit',0)/max(len(reqs),1)*100,1)}%"),
        ("Gap (WRICEF)", by_cls.get("gap", 0), f"{round(by_cls.get('gap',0)/max(len(reqs),1)*100,1)}%"),
        ("", "", ""),
        ("Total WRICEF Items", len(wricef_items), ""),
    ]
    for wtype, cnt in sorted(wricef_type_counts.items()):
        summary_rows.append((f"  {wtype}", cnt, ""))
    summary_rows.append(("Total Config Items", len(config_items), ""))

    start_row = 5
    for i, row_data in enumerate(summary_rows, start=start_row):
        ws1.cell(row=i, column=1).value = row_data[0]
        ws1.cell(row=i, column=2).value = row_data[1]
        ws1.cell(row=i, column=3).value = row_data[2]
        if i == start_row:
            _apply_header_style(ws1, i, 3)

    _auto_width(ws1)

    # ── Tab 2: L3 Process Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("L3 Process Summary")
    headers2 = ["L1 Process", "L2 Process", "L3 Process", "Fit", "Partial Fit", "Gap", "Total", "Gap %"]
    for col, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col).value = h
    _apply_header_style(ws2, 1, len(headers2))

    for row_i, grp in enumerate(sorted(l3_groups.values(), key=lambda x: x["l3"]), start=2):
        total = grp["total"] or 1
        gap_pct = round(grp["gap"] / total * 100, 1)
        ws2.cell(row=row_i, column=1).value = grp["l1"]
        ws2.cell(row=row_i, column=2).value = grp["l2"]
        ws2.cell(row=row_i, column=3).value = grp["l3"]
        ws2.cell(row=row_i, column=4).value = grp["fit"]
        ws2.cell(row=row_i, column=5).value = grp["partial_fit"]
        ws2.cell(row=row_i, column=6).value = grp["gap"]
        ws2.cell(row=row_i, column=7).value = grp["total"]
        ws2.cell(row=row_i, column=8).value = f"{gap_pct}%"
    _auto_width(ws2)

    # ── Tab 3: Requirement Detail ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Requirement Detail")
    headers3 = ["Code", "Title", "Classification", "Priority", "Status", "SAP Module",
                "Moscow Priority", "Description"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col).value = h
    _apply_header_style(ws3, 1, len(headers3))

    for row_i, r in enumerate(reqs, start=2):
        ws3.cell(row=row_i, column=1).value = r.code
        ws3.cell(row=row_i, column=2).value = r.title
        ws3.cell(row=row_i, column=3).value = r.fit_status
        ws3.cell(row=row_i, column=4).value = r.priority
        ws3.cell(row=row_i, column=5).value = r.status
        ws3.cell(row=row_i, column=6).value = r.sap_module
        ws3.cell(row=row_i, column=7).value = getattr(r, "moscow_priority", None)
        ws3.cell(row=row_i, column=8).value = getattr(r, "description", "")
    _auto_width(ws3)

    # ── Tab 4: WRICEF List ────────────────────────────────────────────────────
    if include_wricef:
        ws4 = wb.create_sheet("WRICEF List")
        headers4 = ["Code", "Type", "Title", "Source Req ID", "Priority", "Status",
                    "SAP Module", "Complexity"]
        for col, h in enumerate(headers4, start=1):
            ws4.cell(row=1, column=col).value = h
        _apply_header_style(ws4, 1, len(headers4))

        for row_i, bi in enumerate(wricef_items, start=2):
            ws4.cell(row=row_i, column=1).value = bi.code
            ws4.cell(row=row_i, column=2).value = bi.wricef_type
            ws4.cell(row=row_i, column=3).value = bi.title
            ws4.cell(row=row_i, column=4).value = bi.explore_requirement_id
            ws4.cell(row=row_i, column=5).value = bi.priority
            ws4.cell(row=row_i, column=6).value = bi.status
            ws4.cell(row=row_i, column=7).value = bi.module
            ws4.cell(row=row_i, column=8).value = getattr(bi, "complexity", None)
        _auto_width(ws4)

    # ── Tab 5: Config Items ────────────────────────────────────────────────────
    if include_config:
        ws5 = wb.create_sheet("Config Items")
        headers5 = ["Code", "Title", "Source Req ID", "SAP Module", "IMG Path",
                    "T-Code", "Status"]
        for col, h in enumerate(headers5, start=1):
            ws5.cell(row=1, column=col).value = h
        _apply_header_style(ws5, 1, len(headers5))

        for row_i, ci in enumerate(config_items, start=2):
            ws5.cell(row=row_i, column=1).value = ci.code
            ws5.cell(row=row_i, column=2).value = ci.title
            ws5.cell(row=row_i, column=3).value = ci.explore_requirement_id
            ws5.cell(row=row_i, column=4).value = ci.module
            ws5.cell(row=row_i, column=5).value = getattr(ci, "config_key", None)
            ws5.cell(row=row_i, column=6).value = getattr(ci, "transaction_code", None)
            ws5.cell(row=row_i, column=7).value = ci.status
        _auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_requirement_csv(
    project_id: int,
    tenant_id: int | None,
    workshop_id: int | None = None,
    classification_filter: list[str] | None = None,
) -> str:
    """Generate a CSV containing all ExploreRequirements for a project.

    Fast, simple format for bulk data extraction. Each row is one requirement.
    Tenant isolation enforced via tenant_id scoping (Audit A3).

    Args:
        project_id: Owning project.
        tenant_id: Row-level isolation. None = test environment.
        workshop_id: Optional workshop scope.
        classification_filter: Optional list of fit_status values.

    Returns:
        str: CSV content as a UTF-8 string.
    """
    from app.models.explore import ExploreRequirement

    q = ExploreRequirement.query.filter(
        ExploreRequirement.project_id == project_id,
    )
    if tenant_id is not None:
        q = q.filter(ExploreRequirement.tenant_id == tenant_id)
    if workshop_id is not None:
        q = q.filter(ExploreRequirement.workshop_id == workshop_id)
    if classification_filter:
        q = q.filter(ExploreRequirement.fit_status.in_(classification_filter))

    reqs = q.order_by(ExploreRequirement.code).all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "code", "title", "classification", "priority", "status",
        "sap_module", "moscow_priority", "description",
    ])
    for r in reqs:
        writer.writerow([
            r.code,
            r.title,
            r.fit_status,
            r.priority,
            r.status,
            r.sap_module,
            getattr(r, "moscow_priority", ""),
            (getattr(r, "description", "") or "").replace("\n", " "),
        ])
    return buf.getvalue()
