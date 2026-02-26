"""
SAP Authorization Concept service layer (FDD-I02 / S7-02).

Business logic for:
  - SapAuthRole CRUD (single and composite SAP roles)
  - SapAuthObject management (auth object → field/value assignments)
  - SOD matrix generation and risk acceptance workflow
  - Role → L4 ProcessStep linkage for coverage tracking
  - Authorization concept Excel export (4-sheet workbook)

Architecture (ADR-002):
  This service handles SAP authorization concept data exclusively.
  Platform RBAC (has_permission, require_permission) is managed by
  permission_service.py and must NEVER be called from here.

All functions accept tenant_id as explicit parameter — never read from g.
db.session.commit() is called only inside this service (3-layer contract).
"""

import io
import logging
from datetime import UTC, datetime

from sqlalchemy import select

from app.models import db
from app.models.sap_auth import (
    ROLE_STATUSES,
    ROLE_TYPES,
    SOD_RULES,
    SapAuthObject,
    SapAuthRole,
    SodMatrix,
)

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════════
# Validation helpers
# ═════════════════════════════════════════════════════════════════════════


def _validate_role_name(role_name: str) -> str | None:
    """Return error string if role_name is invalid, else None.

    SAP role naming convention: prefix + module + function, max 30 chars.
    Empty names are always rejected.
    """
    name = role_name.strip() if role_name else ""
    if not name:
        return "role_name is required"
    if len(name) > 30:
        return "role_name must be ≤ 30 characters (SAP PFCG limit)"
    return None


def _require_project(tenant_id: int, project_id: int) -> None:
    """Raise ValueError if tenant has no matching project.

    Why: tenant isolation — a project_id from another tenant must not
    be accessible, and the error must be the same as 'not found'.
    """
    from app.models.program import Program  # local import avoids circular dep

    program = Program.query.filter_by(id=project_id, tenant_id=tenant_id).first()
    if not program:
        raise ValueError(f"Project {project_id} not found in tenant {tenant_id}")


# ═════════════════════════════════════════════════════════════════════════
# SapAuthRole CRUD
# ═════════════════════════════════════════════════════════════════════════


def create_sap_auth_role(tenant_id: int, project_id: int, data: dict) -> dict:
    """Create a new SapAuthRole scoped to the project.

    Business rules:
      - role_name is required (≤ 30 chars, SAP PFCG limit)
      - role_type must be 'single' or 'composite'
      - composite roles must not directly hold auth_objects (they reference
        child single roles via child_role_ids)
      - status defaults to 'draft'

    Args:
        tenant_id:  Owning tenant (isolation scope).
        project_id: Program ID (programs.id — canonical project entity).
        data:       Validated input dict from blueprint.

    Returns:
        Serialized SapAuthRole dict.

    Raises:
        ValueError: If validation fails or project not in tenant.
    """
    _require_project(tenant_id, project_id)

    errors: dict[str, str] = {}

    role_name = data.get("role_name", "")
    name_error = _validate_role_name(role_name)
    if name_error:
        errors["role_name"] = name_error

    role_type = data.get("role_type", "single")
    if role_type not in ROLE_TYPES:
        errors["role_type"] = f"Must be one of: {', '.join(sorted(ROLE_TYPES))}"

    status = data.get("status", "draft")
    if status not in ROLE_STATUSES:
        errors["status"] = f"Must be one of: {', '.join(sorted(ROLE_STATUSES))}"

    if errors:
        raise ValueError(errors)

    role = SapAuthRole(
        tenant_id=tenant_id,
        project_id=project_id,
        role_name=role_name.strip(),
        role_type=role_type,
        description=str(data.get("description", "") or "")[:500],
        sap_module=str(data.get("sap_module", "") or "")[:10] or None,
        org_levels=data.get("org_levels"),
        child_role_ids=data.get("child_role_ids"),
        business_role_description=str(data.get("business_role_description", "") or "")[:200] or None,
        user_count_estimate=data.get("user_count_estimate"),
        linked_process_step_ids=data.get("linked_process_step_ids"),
        status=status,
    )
    db.session.add(role)
    db.session.commit()

    logger.info(
        "SapAuthRole created id=%s tenant=%s project=%s role=%s",
        role.id, tenant_id, project_id, role.role_name,
    )
    return role.to_dict()


def list_sap_auth_roles(tenant_id: int, project_id: int) -> list[dict]:
    """Return all SapAuthRoles for a project, ordered by role_name.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.

    Returns:
        List of serialized role dicts, each including auth_object count.
    """
    stmt = (
        select(SapAuthRole)
        .where(
            SapAuthRole.tenant_id == tenant_id,
            SapAuthRole.project_id == project_id,
        )
        .order_by(SapAuthRole.role_name)
    )
    roles = db.session.execute(stmt).scalars().all()

    result = []
    for role in roles:
        d = role.to_dict()
        d["auth_object_count"] = len(role.auth_objects)
        d["sod_risk_count"] = len(role.sod_risks_as_a) + len(role.sod_risks_as_b)
        result.append(d)
    return result


def get_sap_auth_role(tenant_id: int, project_id: int, role_id: int) -> dict:
    """Return a single SapAuthRole with its auth_objects.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.
        role_id:    SapAuthRole primary key.

    Returns:
        Serialized role dict with nested auth_objects list.

    Raises:
        ValueError: If role not found in this tenant+project.
    """
    role = db.session.get(SapAuthRole, role_id)
    if not role or role.tenant_id != tenant_id or role.project_id != project_id:
        raise ValueError(f"SapAuthRole {role_id} not found")

    d = role.to_dict()
    d["auth_objects"] = [obj.to_dict() for obj in role.auth_objects]
    return d


def update_sap_auth_role(
    tenant_id: int, project_id: int, role_id: int, data: dict
) -> dict:
    """Update mutable fields on a SapAuthRole.

    Only the caller-supplied fields are updated (partial update semantics).
    role_name, role_type, status undergo the same validation as create.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.
        role_id:    Target role.
        data:       Fields to update.

    Returns:
        Updated serialized role dict.

    Raises:
        ValueError: If validation fails or role not found.
    """
    role = db.session.get(SapAuthRole, role_id)
    if not role or role.tenant_id != tenant_id or role.project_id != project_id:
        raise ValueError(f"SapAuthRole {role_id} not found")

    if "role_name" in data:
        err = _validate_role_name(data["role_name"])
        if err:
            raise ValueError({"role_name": err})
        role.role_name = data["role_name"].strip()

    if "role_type" in data:
        if data["role_type"] not in ROLE_TYPES:
            raise ValueError({"role_type": f"Must be one of: {', '.join(sorted(ROLE_TYPES))}"})
        role.role_type = data["role_type"]

    if "status" in data:
        if data["status"] not in ROLE_STATUSES:
            raise ValueError({"status": f"Must be one of: {', '.join(sorted(ROLE_STATUSES))}"})
        role.status = data["status"]

    for field in ("description", "sap_module", "business_role_description",
                  "user_count_estimate", "org_levels", "child_role_ids",
                  "linked_process_step_ids"):
        if field in data:
            setattr(role, field, data[field])

    role.updated_at = datetime.now(UTC)
    db.session.commit()

    logger.info("SapAuthRole updated id=%s tenant=%s", role_id, tenant_id)
    return role.to_dict()


def delete_sap_auth_role(tenant_id: int, project_id: int, role_id: int) -> None:
    """Delete a SapAuthRole and all its auth_objects and SOD rows (cascade).

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.
        role_id:    Target role.

    Raises:
        ValueError: If role not found.
    """
    role = db.session.get(SapAuthRole, role_id)
    if not role or role.tenant_id != tenant_id or role.project_id != project_id:
        raise ValueError(f"SapAuthRole {role_id} not found")

    db.session.delete(role)
    db.session.commit()
    logger.info("SapAuthRole deleted id=%s tenant=%s", role_id, tenant_id)


# ═════════════════════════════════════════════════════════════════════════
# SapAuthObject CRUD
# ═════════════════════════════════════════════════════════════════════════


def add_auth_object(
    tenant_id: int, project_id: int, role_id: int, data: dict
) -> dict:
    """Add a SAP authorization object assignment to a SapAuthRole.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID (used to verify role ownership).
        role_id:    Target SapAuthRole.
        data:       Must contain 'auth_object' (str, ≤10) and
                    'field_values' (dict). Optional: 'auth_object_description',
                    'source'.

    Returns:
        Serialized SapAuthObject dict.

    Raises:
        ValueError: If validation fails or role not found.
    """
    role = db.session.get(SapAuthRole, role_id)
    if not role or role.tenant_id != tenant_id or role.project_id != project_id:
        raise ValueError(f"SapAuthRole {role_id} not found")

    errors: dict[str, str] = {}

    auth_object = str(data.get("auth_object", "")).strip().upper()
    if not auth_object:
        errors["auth_object"] = "auth_object is required"
    elif len(auth_object) > 10:
        errors["auth_object"] = "auth_object must be ≤ 10 characters (SAP object name limit)"

    field_values = data.get("field_values")
    if not isinstance(field_values, dict):
        errors["field_values"] = "field_values must be a dict mapping field names to lists of values"

    if errors:
        raise ValueError(errors)

    obj = SapAuthObject(
        auth_role_id=role_id,
        tenant_id=tenant_id,
        auth_object=auth_object,
        auth_object_description=str(data.get("auth_object_description", "") or "")[:200] or None,
        field_values=field_values,
        source=data.get("source"),
    )
    db.session.add(obj)
    db.session.commit()

    logger.info(
        "SapAuthObject added id=%s role_id=%s object=%s tenant=%s",
        obj.id, role_id, auth_object, tenant_id,
    )
    return obj.to_dict()


def update_auth_object(
    tenant_id: int, role_id: int, obj_id: int, data: dict
) -> dict:
    """Update field_values and/or description on a SapAuthObject.

    Args:
        tenant_id: Tenant isolation scope.
        role_id:   Parent SapAuthRole.
        obj_id:    Target SapAuthObject.
        data:      Fields to update.

    Returns:
        Updated serialized SapAuthObject dict.

    Raises:
        ValueError: If object not found in this tenant+role.
    """
    obj = db.session.get(SapAuthObject, obj_id)
    if not obj or obj.tenant_id != tenant_id or obj.auth_role_id != role_id:
        raise ValueError(f"SapAuthObject {obj_id} not found")

    if "field_values" in data:
        if not isinstance(data["field_values"], dict):
            raise ValueError({"field_values": "Must be a dict"})
        obj.field_values = data["field_values"]

    if "auth_object_description" in data:
        obj.auth_object_description = str(data["auth_object_description"])[:200] or None

    if "source" in data:
        obj.source = data["source"]

    db.session.commit()
    logger.info("SapAuthObject updated id=%s tenant=%s", obj_id, tenant_id)
    return obj.to_dict()


def delete_auth_object(tenant_id: int, role_id: int, obj_id: int) -> None:
    """Remove a SapAuthObject from a role.

    Args:
        tenant_id: Tenant isolation scope.
        role_id:   Parent SapAuthRole.
        obj_id:    Target SapAuthObject.

    Raises:
        ValueError: If object not found.
    """
    obj = db.session.get(SapAuthObject, obj_id)
    if not obj or obj.tenant_id != tenant_id or obj.auth_role_id != role_id:
        raise ValueError(f"SapAuthObject {obj_id} not found")

    db.session.delete(obj)
    db.session.commit()
    logger.info("SapAuthObject deleted id=%s tenant=%s", obj_id, tenant_id)


# ═════════════════════════════════════════════════════════════════════════
# ProcessStep linkage
# ═════════════════════════════════════════════════════════════════════════


def link_role_to_process_steps(
    tenant_id: int, project_id: int, role_id: int, process_step_ids: list[int]
) -> dict:
    """Replace the L4 ProcessStep linkage list on a SapAuthRole.

    Links determine which L4 process steps (user activities) a role covers.
    This drives the coverage report (what % of L4 steps have a role assigned).

    Why replace-all instead of delta: the UI sends the full intended list
    (checkbox state), making idempotent PUT semantics simpler.

    Args:
        tenant_id:        Tenant isolation scope.
        project_id:       Program ID.
        role_id:          Target SapAuthRole.
        process_step_ids: Complete list of L4 ProcessStep IDs to link.

    Returns:
        Updated role dict with linked_process_step_ids populated.

    Raises:
        ValueError: If role not found.
    """
    role = db.session.get(SapAuthRole, role_id)
    if not role or role.tenant_id != tenant_id or role.project_id != project_id:
        raise ValueError(f"SapAuthRole {role_id} not found")

    role.linked_process_step_ids = list(set(process_step_ids))  # dedup
    role.updated_at = datetime.now(UTC)
    db.session.commit()

    logger.info(
        "SapAuthRole %s linked to %d process steps tenant=%s",
        role_id, len(role.linked_process_step_ids), tenant_id,
    )
    return role.to_dict()


# ═════════════════════════════════════════════════════════════════════════
# SOD Matrix generation and risk management
# ═════════════════════════════════════════════════════════════════════════


def generate_sod_matrix(tenant_id: int, project_id: int) -> list[dict]:
    """Detect SOD conflicts between all single-role pairs in the project.

    Algorithm:
      1. Load all single SapAuthRoles for the project.
      2. Load their SapAuthObjects with field_values.
      3. For each role pair (A, B), check every SOD_RULE:
         - If both A and B have the rule's auth_object
         - AND between them they cover ALL conflicting activities
         → insert (or update) a SodMatrix row.
      4. Remove stale SodMatrix rows for role pairs that no longer conflict.

    This is run on-demand (POST /sod-matrix/refresh), not automatically
    on every role change — SOD analysis can be expensive for large projects.

    Activities are compared as sets: if rule says ["01","60"] and
    role A has ["01"] and role B has ["60"] (or A has both) → conflict.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.

    Returns:
        List of SodMatrix dicts (all conflicts for the project, post-refresh).
    """
    # Load all single roles for this project
    stmt = (
        select(SapAuthRole)
        .where(
            SapAuthRole.tenant_id == tenant_id,
            SapAuthRole.project_id == project_id,
            SapAuthRole.role_type == "single",
        )
    )
    roles = db.session.execute(stmt).scalars().all()

    # Build auth-object → activities map per role
    # { role_id: { auth_object_code: set_of_actvt_values } }
    role_objects: dict[int, dict[str, set[str]]] = {}
    for role in roles:
        role_objects[role.id] = {}
        for obj in role.auth_objects:
            actvt_values: list = obj.field_values.get("ACTVT", [])
            role_objects[role.id][obj.auth_object] = {str(v) for v in actvt_values}

    new_conflicts: list[tuple[int, int, str, str, str]] = []  # a_id, b_id, level, desc, ao

    # Check every pair (each pair evaluated once: A < B by ID)
    role_list = list(roles)
    for i, role_a in enumerate(role_list):
        for role_b in role_list[i + 1:]:
            for rule in SOD_RULES:
                ao = rule["auth_object"]
                conflict_acts = set(rule["conflict_activities"])
                acts_a = role_objects.get(role_a.id, {}).get(ao, set())
                acts_b = role_objects.get(role_b.id, {}).get(ao, set())
                combined = acts_a | acts_b
                if conflict_acts.issubset(combined):
                    new_conflicts.append((
                        role_a.id,
                        role_b.id,
                        rule["risk_level"],
                        rule["description"],
                        ao,
                    ))

    # Upsert conflicts — preserve existing is_accepted / mitigating_control
    existing_stmt = select(SodMatrix).where(
        SodMatrix.tenant_id == tenant_id,
        SodMatrix.project_id == project_id,
    )
    existing: list[SodMatrix] = db.session.execute(existing_stmt).scalars().all()
    existing_map: dict[tuple[int, int, str], SodMatrix] = {
        (row.role_a_id, row.role_b_id, row.conflicting_auth_object or ""): row
        for row in existing
    }

    # Remove rows that are no longer conflicting
    new_keys = {(a, b, ao) for a, b, _, _, ao in new_conflicts}
    for key, row in existing_map.items():
        if key not in new_keys:
            db.session.delete(row)

    # Add or update remaining
    for a_id, b_id, risk_level, desc, ao in new_conflicts:
        key = (a_id, b_id, ao)
        if key in existing_map:
            row = existing_map[key]
            row.risk_level = risk_level
            row.risk_description = desc
        else:
            row = SodMatrix(
                tenant_id=tenant_id,
                project_id=project_id,
                role_a_id=a_id,
                role_b_id=b_id,
                risk_level=risk_level,
                risk_description=desc,
                conflicting_auth_object=ao,
            )
            db.session.add(row)

    db.session.commit()

    logger.info(
        "SOD matrix refreshed project=%s tenant=%s conflicts=%d",
        project_id, tenant_id, len(new_conflicts),
    )

    # Return all current SodMatrix rows for this project
    return list_sod_matrix(tenant_id, project_id)


def list_sod_matrix(tenant_id: int, project_id: int) -> list[dict]:
    """Return all SodMatrix rows for a project, enriched with role names.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.

    Returns:
        List of SodMatrix dicts with role_a_name and role_b_name added.
    """
    stmt = (
        select(SodMatrix)
        .where(
            SodMatrix.tenant_id == tenant_id,
            SodMatrix.project_id == project_id,
        )
        .order_by(SodMatrix.risk_level)
    )
    rows = db.session.execute(stmt).scalars().all()

    # Load role name map to avoid N+1
    role_ids = {r.role_a_id for r in rows} | {r.role_b_id for r in rows}
    if role_ids:
        role_stmt = select(SapAuthRole).where(SapAuthRole.id.in_(role_ids))
        role_map = {r.id: r.role_name for r in db.session.execute(role_stmt).scalars().all()}
    else:
        role_map = {}

    result = []
    for row in rows:
        d = row.to_dict()
        d["role_a_name"] = role_map.get(row.role_a_id, "")
        d["role_b_name"] = role_map.get(row.role_b_id, "")
        result.append(d)
    return result


def accept_sod_risk(
    tenant_id: int,
    project_id: int,
    sod_id: int,
    accepted_by_id: int,
    mitigating_control: str,
) -> dict:
    """Mark a SOD risk as accepted with a compensating control description.

    Formal risk acceptance requires:
      - Documenting the mitigating control (e.g. manual approval log)
      - Recording who accepted and when (audit trail)

    Args:
        tenant_id:          Tenant isolation scope.
        project_id:         Program ID.
        sod_id:             SodMatrix primary key.
        accepted_by_id:     User ID of the risk acceptor (audit trail).
        mitigating_control: Non-empty description of the compensating control.

    Returns:
        Updated SodMatrix dict.

    Raises:
        ValueError: If SOD row not found or mitigating_control empty.
    """
    row = db.session.get(SodMatrix, sod_id)
    if not row or row.tenant_id != tenant_id or row.project_id != project_id:
        raise ValueError(f"SodMatrix {sod_id} not found")

    control = str(mitigating_control or "").strip()
    if not control:
        raise ValueError({"mitigating_control": "Required when accepting a SOD risk"})

    row.is_accepted = True
    row.mitigating_control = control[:2000]
    row.accepted_by_id = accepted_by_id
    row.accepted_at = datetime.now(UTC)
    db.session.commit()

    logger.info(
        "SOD risk accepted id=%s by user=%s tenant=%s",
        sod_id, accepted_by_id, tenant_id,
    )
    return row.to_dict()


# ═════════════════════════════════════════════════════════════════════════
# Coverage reporting
# ═════════════════════════════════════════════════════════════════════════


def get_role_coverage(tenant_id: int, project_id: int) -> dict:
    """Calculate what proportion of L4 process steps have a role assigned.

    Coverage = steps that appear in at least one role's linked_process_step_ids
             / total L4 steps in the project.

    This helps the authorization concept team identify which activities
    have no role defined yet.

    Args:
        tenant_id:  Tenant isolation scope.
        project_id: Program ID.

    Returns:
        Dict with total_steps, covered_steps, coverage_pct, missing_step_ids,
        and role_summary list.
    """
    from app.models.explore.process import ProcessStep  # local import
    from app.models.explore.workshop import ExploreWorkshop  # local import

    # ProcessStep has no direct project_id — join through the workshop.
    # ExploreWorkshop.project_id links back to programs.id.
    step_id_stmt = (
        select(ProcessStep.id)
        .join(ExploreWorkshop, ProcessStep.workshop_id == ExploreWorkshop.id)
        .where(
            ExploreWorkshop.tenant_id == tenant_id,
            ExploreWorkshop.project_id == project_id,
        )
    )
    all_step_ids: set = {row[0] for row in db.session.execute(step_id_stmt).all()}

    # Collect all step IDs covered by any role
    role_stmt = select(SapAuthRole).where(
        SapAuthRole.tenant_id == tenant_id,
        SapAuthRole.project_id == project_id,
    )
    roles = db.session.execute(role_stmt).scalars().all()

    covered_ids: set[int] = set()
    role_summary = []
    for role in roles:
        step_ids = set(role.linked_process_step_ids or [])
        covered_ids.update(step_ids)
        role_summary.append({
            "role_id": role.id,
            "role_name": role.role_name,
            "linked_step_count": len(step_ids),
        })

    total = len(all_step_ids)
    # covered = unique step IDs that appear in at least one role's linked list
    # (regardless of whether they exist as ProcessStep rows — teams link steps
    # before workshops are created; counting the intersection would give 0 until then)
    covered = len(covered_ids)
    missing = sorted(all_step_ids - covered_ids)
    # pct denominator: use whichever is larger to avoid > 100%
    denom = max(total, covered)
    pct = round(covered / denom * 100, 1) if denom > 0 else 0.0

    return {
        "total_steps": total,
        "covered_steps": covered,
        "coverage_pct": pct,
        "missing_step_ids": missing,
        "role_summary": role_summary,
    }


# ═════════════════════════════════════════════════════════════════════════
# Excel export
# ═════════════════════════════════════════════════════════════════════════


def export_auth_concept_excel(tenant_id: int, project_id: int) -> bytes:
    """Generate Authorization Concept Excel workbook (4 sheets).

    Sheet 1 — Role List: all SapAuthRoles with status, module, user count
    Sheet 2 — Role-Object Matrix: each role × its auth objects and field values
    Sheet 3 — SOD Matrix: all risk assessments with acceptance status
    Sheet 4 — User Assignment Plan: roles + estimated user counts

    Returns:
        Raw bytes of the .xlsx file for streaming to the client.

    Why openpyxl: already in requirements.txt (used by export_service too).
    weasyprint is excluded (production stability risk — FDD-F03 §review).
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    CENTER = Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    def style_header_row(ws, headers: list[str]) -> None:
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # ── Sheet 1: Role List ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Role List"
    style_header_row(ws1, [
        "Role Name", "Type", "SAP Module", "Status",
        "Business Description", "Est. Users", "Objects #",
    ])

    roles = list_sap_auth_roles(tenant_id, project_id)
    for r in roles:
        ws1.append([
            r["role_name"],
            r["role_type"],
            r.get("sap_module") or "",
            r["status"],
            r.get("business_role_description") or "",
            r.get("user_count_estimate") or "",
            r.get("auth_object_count", 0),
        ])

    # ── Sheet 2: Role-Object Matrix ───────────────────────────────────────
    ws2 = wb.create_sheet("Role-Object Matrix")
    style_header_row(ws2, [
        "Role Name", "Auth Object", "Description",
        "Field Values (JSON)", "Source",
    ])

    for r in roles:
        role_detail = get_sap_auth_role(tenant_id, project_id, r["id"])
        for obj in role_detail.get("auth_objects", []):
            import json
            ws2.append([
                r["role_name"],
                obj["auth_object"],
                obj.get("auth_object_description") or "",
                json.dumps(obj.get("field_values", {})),
                obj.get("source") or "",
            ])

    # ── Sheet 3: SOD Matrix ───────────────────────────────────────────────
    ws3 = wb.create_sheet("SOD Matrix")
    style_header_row(ws3, [
        "Role A", "Role B", "Auth Object", "Risk Level",
        "Description", "Mitigating Control", "Accepted",
    ])

    sod_rows = list_sod_matrix(tenant_id, project_id)
    for row in sod_rows:
        ws3.append([
            row.get("role_a_name", ""),
            row.get("role_b_name", ""),
            row.get("conflicting_auth_object") or "",
            row["risk_level"],
            row.get("risk_description") or "",
            row.get("mitigating_control") or "",
            "Yes" if row.get("is_accepted") else "No",
        ])

    # ── Sheet 4: User Assignment Plan ─────────────────────────────────────
    ws4 = wb.create_sheet("User Assignment Plan")
    style_header_row(ws4, [
        "Role Name", "Type", "Module", "Est. Users",
        "Business Function", "Status",
    ])

    for r in roles:
        ws4.append([
            r["role_name"],
            r["role_type"],
            r.get("sap_module") or "",
            r.get("user_count_estimate") or "",
            r.get("business_role_description") or "",
            r["status"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
